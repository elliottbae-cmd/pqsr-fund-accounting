"""Generate the updated Excel accounting workbook."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from config.fund_config import FUND_NAME, INVESTORS, FIXED_ASSETS


HEADER_FILL = PatternFill("solid", fgColor="F4A523")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
NUM_FMT = '#,##0.00'
ACCT_FMT = '$#,##0'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _write_header(ws, row, col, text, width=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    if width:
        ws.column_dimensions[cell.column_letter].width = width


def _write_row(ws, row, label, value, label_col=1, val_col=2, bold=False):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = BOLD_FONT if bold else LABEL_FONT
    vc = ws.cell(row=row, column=val_col, value=value)
    vc.font = BOLD_FONT if bold else LABEL_FONT
    vc.number_format = NUM_FMT
    vc.alignment = Alignment(horizontal="right")


def generate_excel_workbook(bs, is_accounts, cash_flow, totals, distribution_data,
                            journal_entries, as_of_date):
    """Generate an Excel workbook with BS, IS, AJEs, and distributions."""
    buffer = io.BytesIO()
    wb = Workbook()

    # --- Balance Sheet ---
    ws_bs = wb.active
    ws_bs.title = "Balance Sheet"
    ws_bs.column_dimensions['A'].width = 40
    ws_bs.column_dimensions['B'].width = 18

    ws_bs.cell(row=1, column=1, value=FUND_NAME).font = BOLD_FONT
    ws_bs.cell(row=2, column=1, value="Balance Sheet").font = BOLD_FONT
    ws_bs.cell(row=3, column=1, value=as_of_date.strftime("%m/%d/%Y")).font = LABEL_FONT

    r = 5
    _write_header(ws_bs, r, 1, "ASSETS")
    _write_header(ws_bs, r, 2, "")
    r += 1
    _write_row(ws_bs, r, "Cash", bs["Cash"]); r += 1
    r += 1
    ws_bs.cell(row=r, column=1, value="FIXED ASSETS").font = BOLD_FONT; r += 1

    for asset in ["Land", "Building", "Land Improvements", "F&F", "Equipment", "Signage"]:
        label = "Furniture & Fixtures" if asset == "F&F" else asset
        _write_row(ws_bs, r, label, bs[asset]); r += 1

    for asset in ["Building", "Land Improvements", "F&F", "Equipment", "Signage"]:
        label = f"{'F&F' if asset == 'F&F' else asset} - Accum. Depreciation"
        _write_row(ws_bs, r, label, bs[f"{asset} A/D"]); r += 1

    _write_row(ws_bs, r, "Total Fixed Assets (Net)", totals["total_fa_net"], bold=True); r += 1
    r += 1
    ws_bs.cell(row=r, column=1, value="OTHER ASSETS").font = BOLD_FONT; r += 1
    _write_row(ws_bs, r, "Capitalized Origination Fee", bs["Capitalized Origination Fee"]); r += 1
    _write_row(ws_bs, r, "Accumulated Amortization", bs["Accumulated Amortization"]); r += 1
    _write_row(ws_bs, r, "Total Other Assets", totals["total_other_assets"], bold=True); r += 1
    r += 1
    _write_row(ws_bs, r, "Total Assets", totals["total_assets"], bold=True); r += 2

    _write_header(ws_bs, r, 1, "LIABILITIES")
    _write_header(ws_bs, r, 2, ""); r += 1
    _write_row(ws_bs, r, "Note Payable - BBV", bs["Note Payable - BBV"]); r += 1
    _write_row(ws_bs, r, "Due to PSP Investments, LLC", bs["Due to PSP Investments, LLC"]); r += 1
    _write_row(ws_bs, r, "Deferred Rental Revenue", bs["Deferred Rental Revenue"]); r += 1
    _write_row(ws_bs, r, "Total Liabilities", totals["total_liabilities"], bold=True); r += 2

    ws_bs.cell(row=r, column=1, value="MEMBERS' EQUITY").font = BOLD_FONT; r += 1
    for inv_key in INVESTORS:
        _write_row(ws_bs, r, f"Contributions - {inv_key}", bs[f"Contributions - {inv_key}"]); r += 1
    for inv_key in INVESTORS:
        _write_row(ws_bs, r, f"Distributions - {inv_key}", bs[f"Distributions - {inv_key}"]); r += 1
    _write_row(ws_bs, r, "CY Net Income", bs["CY Net Income"]); r += 1
    _write_row(ws_bs, r, "Retained Earnings", bs["Retained Earnings"]); r += 1
    _write_row(ws_bs, r, "Total Equity", totals["total_equity"], bold=True); r += 1
    r += 1
    _write_row(ws_bs, r, "Total Liabilities / Equity", totals["total_liabilities_equity"], bold=True)

    # --- Income Statement ---
    ws_is = wb.create_sheet("Income Statement")
    ws_is.column_dimensions['A'].width = 40
    ws_is.column_dimensions['B'].width = 18

    ws_is.cell(row=1, column=1, value=FUND_NAME).font = BOLD_FONT
    ws_is.cell(row=2, column=1, value="Income Statement").font = BOLD_FONT
    ws_is.cell(row=3, column=1, value=as_of_date.strftime("%m/%d/%Y")).font = LABEL_FONT

    r = 5
    _write_header(ws_is, r, 1, "REVENUE"); _write_header(ws_is, r, 2, ""); r += 1
    _write_row(ws_is, r, "Rental Income", is_accounts["Rental Income"]); r += 2

    ws_is.cell(row=r, column=1, value="EXPENSES").font = BOLD_FONT; r += 1
    for exp in ["Interest Expense", "Accounting & Tax Fees", "Bank Fees", "Depreciation Expense"]:
        val = is_accounts.get(exp, 0)
        _write_row(ws_is, r, exp, val if val else 0); r += 1

    total_exp = sum(is_accounts[k] for k in is_accounts if k != "Rental Income")
    _write_row(ws_is, r, "Total Expenses", total_exp, bold=True); r += 1
    net_inc = is_accounts["Rental Income"] - total_exp
    _write_row(ws_is, r, "Net Income", net_inc, bold=True); r += 2

    _write_header(ws_is, r, 1, "CASH FLOW METRICS"); _write_header(ws_is, r, 2, ""); r += 1
    _write_row(ws_is, r, "EBITDA", cash_flow["EBITDA"]); r += 1
    _write_row(ws_is, r, "Less: Interest Expense", -cash_flow["Interest Expense"]); r += 1
    _write_row(ws_is, r, "Less: Principal Payments", -cash_flow["Principal Payments"]); r += 1
    _write_row(ws_is, r, "Free Cash Flow (FCF)", cash_flow["FCF"], bold=True); r += 2
    _write_row(ws_is, r, "DSCR", cash_flow["DSCR"]); r += 1
    ws_is.cell(row=r-1, column=2).number_format = '0.00"x"'

    # --- Journal Entries ---
    ws_aje = wb.create_sheet("Journal Entries")
    ws_aje.column_dimensions['A'].width = 12
    ws_aje.column_dimensions['B'].width = 35
    ws_aje.column_dimensions['C'].width = 15
    ws_aje.column_dimensions['D'].width = 15

    _write_header(ws_aje, 1, 1, "Date")
    _write_header(ws_aje, 1, 2, "Account")
    _write_header(ws_aje, 1, 3, "Debit")
    _write_header(ws_aje, 1, 4, "Credit")

    r = 2
    for entry in journal_entries:
        ws_aje.cell(row=r, column=1, value=entry["date"].strftime("%m/%d/%Y")).font = LABEL_FONT
        ws_aje.cell(row=r, column=2, value=entry["description"]).font = BOLD_FONT
        r += 1
        for acct, amt in entry["debits"].items():
            ws_aje.cell(row=r, column=2, value=f"  {acct}").font = LABEL_FONT
            ws_aje.cell(row=r, column=3, value=amt).number_format = NUM_FMT
            r += 1
        for acct, amt in entry["credits"].items():
            ws_aje.cell(row=r, column=2, value=f"      {acct}").font = LABEL_FONT
            ws_aje.cell(row=r, column=4, value=amt).number_format = NUM_FMT
            r += 1
        r += 1

    wb.save(buffer)
    buffer.seek(0)
    return buffer
