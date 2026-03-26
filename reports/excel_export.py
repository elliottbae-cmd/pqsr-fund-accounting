"""Shared Excel export utility with professional formatting for all financial pages."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Brand colors
GOLD_HEX = "F4A523"
GOLD_DARK_HEX = "C78A1E"
WHITE_HEX = "FFFFFF"
LIGHT_GRAY_HEX = "F5F5F5"
ALT_ROW_HEX = "F9F9F9"
TEXT_DARK_HEX = "494949"
BORDER_HEX = "D0D0D0"

# Styles
HEADER_FILL = PatternFill("solid", fgColor=GOLD_HEX)
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE_HEX, size=10)
SECTION_FILL = PatternFill("solid", fgColor=GOLD_DARK_HEX)
SECTION_FONT = Font(name="Calibri", bold=True, color=WHITE_HEX, size=10)
LABEL_FONT = Font(name="Calibri", size=10, color=TEXT_DARK_HEX)
BOLD_FONT = Font(name="Calibri", bold=True, size=10, color=TEXT_DARK_HEX)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=TEXT_DARK_HEX)
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=11, color="797979")
NUM_FMT = '#,##0.00'
NUM_FMT_WHOLE = '#,##0'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_HEX),
    right=Side(style="thin", color=BORDER_HEX),
    top=Side(style="thin", color=BORDER_HEX),
    bottom=Side(style="thin", color=BORDER_HEX),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color=GOLD_HEX),
)


def _apply_header_row(ws, row, num_cols):
    """Style a row as a gold header."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _apply_section_row(ws, row, num_cols):
    """Style a row as a darker gold section header."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = THIN_BORDER


def _apply_data_row(ws, row, num_cols, bold=False, alt=False):
    """Style a data row with optional bold and alternating background."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_FONT if bold else LABEL_FONT
        cell.border = THIN_BORDER
        if alt:
            cell.fill = PatternFill("solid", fgColor=ALT_ROW_HEX)


def _apply_total_row(ws, row, num_cols):
    """Style a totals row with bold and top border."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_FONT
        cell.border = Border(
            top=Side(style="medium", color=GOLD_HEX),
            bottom=Side(style="double", color=GOLD_HEX),
            left=Side(style="thin", color=BORDER_HEX),
            right=Side(style="thin", color=BORDER_HEX),
        )


def _write_title(ws, row, title, subtitle=None, num_cols=2):
    """Write a title and subtitle at the top of a sheet."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=num_cols
    )
    if subtitle:
        cell2 = ws.cell(row=row + 1, column=1, value=subtitle)
        cell2.font = SUBTITLE_FONT
        ws.merge_cells(
            start_row=row + 1, start_column=1, end_row=row + 1, end_column=num_cols
        )
        # Gold underline bar
        for c in range(1, num_cols + 1):
            ws.cell(row=row + 2, column=c).border = BOTTOM_BORDER
        return row + 3
    for c in range(1, num_cols + 1):
        ws.cell(row=row + 1, column=c).border = BOTTOM_BORDER
    return row + 2


def _auto_width(ws, min_width=12, max_width=22):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = min(cell_len + 2, max_width)
        ws.column_dimensions[col_letter].width = max_len


def _set_num_format(ws, row, col, value, fmt=NUM_FMT):
    """Write a numeric value with formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if isinstance(value, (int, float)):
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    return cell


# ==================== EXPORT FUNCTIONS ====================

def export_current_financials(bs, is_accounts, cash_flow, totals, as_of_date, fund_name, investors):
    """Export Financials - Current as a professionally formatted Excel workbook."""
    buffer = io.BytesIO()
    wb = Workbook()

    # --- Balance Sheet ---
    ws_bs = wb.active
    ws_bs.title = "Balance Sheet"
    num_cols = 2
    r = _write_title(
        ws_bs, 1,
        "{} | Balance Sheet".format(fund_name),
        as_of_date.strftime("%m/%d/%Y"),
        num_cols,
    )
    r += 1

    # Headers
    ws_bs.cell(row=r, column=1, value="Account")
    ws_bs.cell(row=r, column=2, value="Amount")
    _apply_header_row(ws_bs, r, num_cols)
    r += 1

    def _bs_row(label, value, section=False, total=False):
        nonlocal r
        ws_bs.cell(row=r, column=1, value=label)
        if value is not None:
            _set_num_format(ws_bs, r, 2, value)
        if section:
            _apply_section_row(ws_bs, r, num_cols)
        elif total:
            _apply_total_row(ws_bs, r, num_cols)
        else:
            _apply_data_row(ws_bs, r, num_cols, alt=(r % 2 == 0))
        r += 1

    _bs_row("ASSETS", None, section=True)
    _bs_row("Cash", bs.get("Cash", 0))

    _bs_row("FIXED ASSETS", None, section=True)
    for asset in ["Land", "Building", "Land Improvements", "F&F", "Equipment", "Signage"]:
        label = "Furniture & Fixtures" if asset == "F&F" else asset
        _bs_row(label, bs.get(asset, 0))
    for asset in ["Building", "Land Improvements", "F&F", "Equipment", "Signage"]:
        _bs_row("{} - Accum. Depreciation".format(asset), bs.get("{} A/D".format(asset), 0))
    _bs_row("Total Fixed Assets (Net)", totals.get("total_fa_net", 0), total=True)

    _bs_row("OTHER ASSETS", None, section=True)
    _bs_row("Capitalized Origination Fee", bs.get("Capitalized Origination Fee", 0))
    _bs_row("Accumulated Amortization", bs.get("Accumulated Amortization", 0))
    _bs_row("Total Other Assets", totals.get("total_other_assets", 0), total=True)

    _bs_row("Total Assets", totals.get("total_assets", 0), total=True)
    r += 1

    _bs_row("LIABILITIES", None, section=True)
    _bs_row("Note Payable - BBV", bs.get("Note Payable - BBV", 0))
    _bs_row("Due to PSP Investments, LLC", bs.get("Due to PSP Investments, LLC", 0))
    _bs_row("Deferred Rental Revenue", bs.get("Deferred Rental Revenue", 0))
    _bs_row("Total Liabilities", totals.get("total_liabilities", 0), total=True)
    r += 1

    _bs_row("MEMBERS' EQUITY", None, section=True)
    for inv_key in investors:
        _bs_row("Contributions - {}".format(inv_key), bs.get("Contributions - {}".format(inv_key), 0))
    for inv_key in investors:
        _bs_row("Distributions - {}".format(inv_key), bs.get("Distributions - {}".format(inv_key), 0))
    _bs_row("CY Net Income", bs.get("CY Net Income", 0))
    _bs_row("Retained Earnings", bs.get("Retained Earnings", 0))
    _bs_row("Total Equity", totals.get("total_equity", 0), total=True)
    r += 1
    _bs_row("Total Liabilities / Equity", totals.get("total_liabilities_equity", 0), total=True)

    ws_bs.column_dimensions["A"].width = 38
    ws_bs.column_dimensions["B"].width = 20

    # --- Income Statement ---
    ws_is = wb.create_sheet("Income Statement")
    r = _write_title(
        ws_is, 1,
        "{} | Income Statement".format(fund_name),
        as_of_date.strftime("%m/%d/%Y"),
        2,
    )
    r += 1
    ws_is.cell(row=r, column=1, value="Account")
    ws_is.cell(row=r, column=2, value="Amount")
    _apply_header_row(ws_is, r, 2)
    r += 1

    def _is_row(label, value, section=False, total=False):
        nonlocal r
        ws_is.cell(row=r, column=1, value=label)
        if value is not None:
            _set_num_format(ws_is, r, 2, value)
        if section:
            _apply_section_row(ws_is, r, 2)
        elif total:
            _apply_total_row(ws_is, r, 2)
        else:
            _apply_data_row(ws_is, r, 2, alt=(r % 2 == 0))
        r += 1

    _is_row("REVENUE", None, section=True)
    _is_row("Rental Income", is_accounts.get("Rental Income", 0))
    r += 1
    _is_row("EXPENSES", None, section=True)
    for exp in ["Interest Expense", "Accounting & Tax Fees", "Bank Fees",
                "Appraisals", "Taxes & Licenses", "Survey Fees",
                "Origination Fee - Amort", "Depreciation Expense"]:
        val = is_accounts.get(exp, 0)
        if val:
            _is_row(exp, val)
    total_exp = sum(is_accounts.get(k, 0) for k in is_accounts if k != "Rental Income")
    _is_row("Total Expenses", total_exp, total=True)
    net_inc = is_accounts.get("Rental Income", 0) - total_exp
    _is_row("Net Income", net_inc, total=True)

    r += 2
    _is_row("CASH FLOW METRICS", None, section=True)
    _is_row("EBITDA", cash_flow.get("EBITDA", 0))
    _is_row("Less: Interest Expense", -cash_flow.get("Interest Expense", 0))
    _is_row("Less: Principal Payments", -cash_flow.get("Principal Payments", 0))
    _is_row("Free Cash Flow (FCF)", cash_flow.get("FCF", 0), total=True)
    r += 1
    ws_is.cell(row=r, column=1, value="DSCR").font = BOLD_FONT
    ws_is.cell(row=r, column=2, value=cash_flow.get("DSCR", 0))
    ws_is.cell(row=r, column=2).number_format = '0.0000"x"'
    ws_is.cell(row=r, column=2).font = BOLD_FONT

    ws_is.column_dimensions["A"].width = 38
    ws_is.column_dimensions["B"].width = 20

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_monthly_financials(
    all_bs, all_is, all_cf, all_totals,
    year_periods, month_labels, monthly_is_deltas,
    baseline_bs, baseline_is, fund_name, investors,
):
    """Export Financials - Monthly with side-by-side month columns."""
    buffer = io.BytesIO()
    wb = Workbook()
    num_months = len(year_periods)
    num_cols = num_months + 2  # Account + Baseline + months

    # --- Balance Sheet ---
    ws_bs = wb.active
    ws_bs.title = "BS - Monthly"
    r = _write_title(ws_bs, 1, "{} | Balance Sheet - Monthly".format(fund_name), None, num_cols)
    r += 1

    # Headers
    ws_bs.cell(row=r, column=1, value="Account")
    ws_bs.cell(row=r, column=2, value="12/31/2025")
    for i, ml in enumerate(month_labels):
        ws_bs.cell(row=r, column=i + 3, value=ml)
    _apply_header_row(ws_bs, r, num_cols)
    r += 1

    bs_accounts = [
        ("ASSETS", None, True),
        ("Cash", "Cash", False),
        ("FIXED ASSETS", None, True),
        ("Land", "Land", False),
        ("Building", "Building", False),
        ("Land Improvements", "Land Improvements", False),
        ("Furniture & Fixtures", "F&F", False),
        ("Equipment", "Equipment", False),
        ("Signage", "Signage", False),
        ("Building - A/D", "Building A/D", False),
        ("Land Improvements - A/D", "Land Improvements A/D", False),
        ("F&F - A/D", "F&F A/D", False),
        ("Equipment - A/D", "Equipment A/D", False),
        ("Signage - A/D", "Signage A/D", False),
        ("OTHER ASSETS", None, True),
        ("Capitalized Origination Fee", "Capitalized Origination Fee", False),
        ("Accumulated Amortization", "Accumulated Amortization", False),
        ("LIABILITIES", None, True),
        ("Note Payable - BBV", "Note Payable - BBV", False),
        ("Due to PSP Investments, LLC", "Due to PSP Investments, LLC", False),
        ("MEMBERS' EQUITY", None, True),
    ]
    for inv_key in investors:
        bs_accounts.append(("Contributions - {}".format(inv_key), "Contributions - {}".format(inv_key), False))
    for inv_key in investors:
        bs_accounts.append(("Distributions - {}".format(inv_key), "Distributions - {}".format(inv_key), False))
    bs_accounts.extend([
        ("CY Net Income", "CY Net Income", False),
        ("Retained Earnings", "Retained Earnings", False),
    ])

    for label, acct_key, is_section in bs_accounts:
        ws_bs.cell(row=r, column=1, value=label)
        if acct_key:
            _set_num_format(ws_bs, r, 2, baseline_bs.get(acct_key, 0))
            for i, pk in enumerate(year_periods):
                _set_num_format(ws_bs, r, i + 3, all_bs.get(pk, {}).get(acct_key, 0))
        if is_section:
            _apply_section_row(ws_bs, r, num_cols)
        else:
            _apply_data_row(ws_bs, r, num_cols, alt=(r % 2 == 0))
        r += 1

    # Totals
    for total_label, total_key in [
        ("Total Assets", "total_assets"),
        ("Total Liabilities", "total_liabilities"),
        ("Total Equity", "total_equity"),
        ("Total L + E", "total_liabilities_equity"),
    ]:
        ws_bs.cell(row=r, column=1, value=total_label)
        for i, pk in enumerate(year_periods):
            _set_num_format(ws_bs, r, i + 3, all_totals.get(pk, {}).get(total_key, 0))
        _apply_total_row(ws_bs, r, num_cols)
        r += 1

    ws_bs.column_dimensions["A"].width = 35
    for c in range(2, num_cols + 1):
        ws_bs.column_dimensions[get_column_letter(c)].width = 16

    # --- Income Statement ---
    ws_is = wb.create_sheet("IS - Monthly")
    r = _write_title(ws_is, 1, "{} | Income Statement - Monthly Activity".format(fund_name), None, num_months + 1)
    r += 1

    ws_is.cell(row=r, column=1, value="Account")
    for i, ml in enumerate(month_labels):
        ws_is.cell(row=r, column=i + 2, value=ml)
    _apply_header_row(ws_is, r, num_months + 1)
    r += 1

    is_order = [
        ("REVENUE", None, True),
        ("Rental Income", "Rental Income", False),
        ("EXPENSES", None, True),
        ("Interest Expense", "Interest Expense", False),
        ("Accounting & Tax Fees", "Accounting & Tax Fees", False),
        ("Bank Fees", "Bank Fees", False),
        ("Depreciation Expense", "Depreciation Expense", False),
    ]

    for label, acct_key, is_section in is_order:
        ws_is.cell(row=r, column=1, value=label)
        if acct_key:
            for i in range(num_months):
                _set_num_format(ws_is, r, i + 2, monthly_is_deltas[i].get(acct_key, 0))
        if is_section:
            _apply_section_row(ws_is, r, num_months + 1)
        else:
            _apply_data_row(ws_is, r, num_months + 1, alt=(r % 2 == 0))
        r += 1

    # Total Expenses
    ws_is.cell(row=r, column=1, value="Total Expenses")
    for i in range(num_months):
        total_exp = sum(monthly_is_deltas[i].get(k, 0) for k in monthly_is_deltas[i] if k != "Rental Income")
        _set_num_format(ws_is, r, i + 2, total_exp)
    _apply_total_row(ws_is, r, num_months + 1)
    r += 1

    # Net Income
    ws_is.cell(row=r, column=1, value="Net Income")
    for i in range(num_months):
        revenue = monthly_is_deltas[i].get("Rental Income", 0)
        total_exp = sum(monthly_is_deltas[i].get(k, 0) for k in monthly_is_deltas[i] if k != "Rental Income")
        _set_num_format(ws_is, r, i + 2, revenue - total_exp)
    _apply_total_row(ws_is, r, num_months + 1)

    ws_is.column_dimensions["A"].width = 30
    for c in range(2, num_months + 2):
        ws_is.column_dimensions[get_column_letter(c)].width = 16

    # --- Cash Flow ---
    ws_cf = wb.create_sheet("CF - Monthly")
    r = _write_title(ws_cf, 1, "{} | Cash Flow - Monthly".format(fund_name), None, num_months + 1)
    r += 1

    ws_cf.cell(row=r, column=1, value="Metric")
    for i, ml in enumerate(month_labels):
        ws_cf.cell(row=r, column=i + 2, value=ml)
    _apply_header_row(ws_cf, r, num_months + 1)
    r += 1

    for metric in ["EBITDA", "Interest Expense", "Principal Payments", "FCF", "DSCR"]:
        ws_cf.cell(row=r, column=1, value=metric)
        for i, pk in enumerate(year_periods):
            val = all_cf.get(pk, {}).get(metric, 0)
            if metric == "DSCR":
                cell = ws_cf.cell(row=r, column=i + 2, value=val)
                cell.number_format = '0.0000"x"'
            else:
                _set_num_format(ws_cf, r, i + 2, val)
        is_total = metric in ("FCF",)
        if is_total:
            _apply_total_row(ws_cf, r, num_months + 1)
        else:
            _apply_data_row(ws_cf, r, num_months + 1, alt=(r % 2 == 0))
        r += 1

    ws_cf.column_dimensions["A"].width = 25
    for c in range(2, num_months + 2):
        ws_cf.column_dimensions[get_column_letter(c)].width = 16

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_quarterly_financials(
    quarter_exports, fund_name, selected_year,
):
    """Export Financials - Quarterly with a sheet per quarter.

    Args:
        quarter_exports: list of dicts, one per quarter, each containing:
            q_num, months_posted, is_complete,
            bs_end, totals_end, is_delta, cf_end,
            prior_label, end_label, bs_prior
        fund_name: fund name string
        selected_year: year int
    """
    buffer = io.BytesIO()
    wb = Workbook()
    first_sheet = True

    for qe in quarter_exports:
        q_num = qe["q_num"]
        sheet_name = "Q{}".format(q_num)
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        r = _write_title(
            ws, 1,
            "{} | Q{} {} Financials".format(fund_name, q_num, selected_year),
            "{} ({})".format(
                "Complete" if qe["is_complete"] else "Partial",
                ", ".join(qe["months_posted"]),
            ),
            4,
        )
        r += 1

        # BS section
        ws.cell(row=r, column=1, value="BALANCE SHEET")
        _apply_section_row(ws, r, 4)
        r += 1
        ws.cell(row=r, column=1, value="Account")
        ws.cell(row=r, column=2, value=qe["prior_label"])
        ws.cell(row=r, column=3, value=qe["end_label"])
        ws.cell(row=r, column=4, value="Change")
        _apply_header_row(ws, r, 4)
        r += 1

        bs_end = qe["bs_end"]
        bs_prior = qe["bs_prior"]
        for label, acct_key in [
            ("Cash", "Cash"), ("Land", "Land"), ("Building", "Building"),
            ("Land Improvements", "Land Improvements"),
            ("F&F", "F&F"), ("Equipment", "Equipment"), ("Signage", "Signage"),
            ("Note Payable - BBV", "Note Payable - BBV"),
            ("CY Net Income", "CY Net Income"),
        ]:
            prior_val = bs_prior.get(acct_key, 0)
            end_val = bs_end.get(acct_key, 0)
            ws.cell(row=r, column=1, value=label)
            _set_num_format(ws, r, 2, prior_val)
            _set_num_format(ws, r, 3, end_val)
            _set_num_format(ws, r, 4, end_val - prior_val)
            _apply_data_row(ws, r, 4, alt=(r % 2 == 0))
            r += 1

        # Totals
        for total_label, total_key in [
            ("Total Assets", "total_assets"),
            ("Total Equity", "total_equity"),
        ]:
            ws.cell(row=r, column=1, value=total_label)
            _set_num_format(ws, r, 3, qe["totals_end"].get(total_key, 0))
            _apply_total_row(ws, r, 4)
            r += 1

        r += 1

        # IS section
        ws.cell(row=r, column=1, value="INCOME STATEMENT - Q{} ACTIVITY".format(q_num))
        _apply_section_row(ws, r, 4)
        r += 1
        ws.cell(row=r, column=1, value="Account")
        ws.cell(row=r, column=2, value="Amount")
        _apply_header_row(ws, r, 2)
        r += 1

        is_delta = qe["is_delta"]
        for acct in ["Rental Income", "Interest Expense", "Accounting & Tax Fees",
                      "Bank Fees", "Depreciation Expense"]:
            val = is_delta.get(acct, 0)
            if val:
                ws.cell(row=r, column=1, value=acct)
                _set_num_format(ws, r, 2, val)
                _apply_data_row(ws, r, 2, alt=(r % 2 == 0))
                r += 1

        total_exp = sum(is_delta.get(k, 0) for k in is_delta if k != "Rental Income")
        ws.cell(row=r, column=1, value="Total Expenses")
        _set_num_format(ws, r, 2, total_exp)
        _apply_total_row(ws, r, 2)
        r += 1
        ws.cell(row=r, column=1, value="Net Income")
        _set_num_format(ws, r, 2, is_delta.get("Rental Income", 0) - total_exp)
        _apply_total_row(ws, r, 2)
        r += 2

        # Cash Flow
        ws.cell(row=r, column=1, value="CASH FLOW")
        _apply_section_row(ws, r, 2)
        r += 1
        cf_end = qe["cf_end"]
        for metric in ["EBITDA", "Interest Expense", "Principal Payments", "FCF"]:
            ws.cell(row=r, column=1, value=metric)
            _set_num_format(ws, r, 2, cf_end.get(metric, 0))
            _apply_data_row(ws, r, 2, bold=(metric == "FCF"))
            r += 1
        ws.cell(row=r, column=1, value="DSCR").font = BOLD_FONT
        ws.cell(row=r, column=2, value=cf_end.get("DSCR", 0))
        ws.cell(row=r, column=2).number_format = '0.0000"x"'

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_financial_history(
    bs, is_accounts, cf, totals, ajes, txns,
    baseline_bs, baseline_is, baseline_cf,
    selected_end, fund_name, investors, investor_report_names,
    distribution_history, db_dists, amort_schedule, loan_balance,
):
    """Export Financial History page — all tabs in one workbook."""
    buffer = io.BytesIO()
    wb = Workbook()

    end_label = selected_end.strftime("%m/%d/%Y")

    # --- BS ---
    ws_bs = wb.active
    ws_bs.title = "Balance Sheet"
    r = _write_title(ws_bs, 1, "{} | Balance Sheet".format(fund_name), end_label, 4)
    r += 1

    ws_bs.cell(row=r, column=1, value="Account")
    ws_bs.cell(row=r, column=2, value="12/31/2025")
    ws_bs.cell(row=r, column=3, value=end_label)
    ws_bs.cell(row=r, column=4, value="Change")
    _apply_header_row(ws_bs, r, 4)
    r += 1

    bs_accounts = [
        ("Cash", "Cash"), ("Land", "Land"), ("Building", "Building"),
        ("Land Improvements", "Land Improvements"),
        ("Furniture & Fixtures", "F&F"), ("Equipment", "Equipment"),
        ("Signage", "Signage"),
        ("Building - A/D", "Building A/D"),
        ("Land Improvements - A/D", "Land Improvements A/D"),
        ("F&F - A/D", "F&F A/D"),
        ("Equipment - A/D", "Equipment A/D"),
        ("Signage - A/D", "Signage A/D"),
        ("Capitalized Origination Fee", "Capitalized Origination Fee"),
        ("Accumulated Amortization", "Accumulated Amortization"),
        ("Note Payable - BBV", "Note Payable - BBV"),
        ("Due to PSP Investments, LLC", "Due to PSP Investments, LLC"),
        ("CY Net Income", "CY Net Income"),
        ("Retained Earnings", "Retained Earnings"),
    ]
    for inv_key in investors:
        bs_accounts.append(("Contributions - {}".format(inv_key), "Contributions - {}".format(inv_key)))
    for inv_key in investors:
        bs_accounts.append(("Distributions - {}".format(inv_key), "Distributions - {}".format(inv_key)))

    for label, acct_key in bs_accounts:
        beg = baseline_bs.get(acct_key, 0)
        end = bs.get(acct_key, 0)
        ws_bs.cell(row=r, column=1, value=label)
        _set_num_format(ws_bs, r, 2, beg)
        _set_num_format(ws_bs, r, 3, end)
        _set_num_format(ws_bs, r, 4, end - beg)
        _apply_data_row(ws_bs, r, 4, alt=(r % 2 == 0))
        r += 1

    for total_label, total_key in [
        ("Total Assets", "total_assets"),
        ("Total Liabilities", "total_liabilities"),
        ("Total Equity", "total_equity"),
        ("Total Liabilities / Equity", "total_liabilities_equity"),
    ]:
        ws_bs.cell(row=r, column=1, value=total_label)
        _set_num_format(ws_bs, r, 3, totals.get(total_key, 0))
        _apply_total_row(ws_bs, r, 4)
        r += 1

    ws_bs.column_dimensions["A"].width = 38
    for c in ["B", "C", "D"]:
        ws_bs.column_dimensions[c].width = 18

    # --- IS ---
    ws_is = wb.create_sheet("Income Statement")
    r = _write_title(ws_is, 1, "{} | Income Statement".format(fund_name), end_label, 4)
    r += 1
    ws_is.cell(row=r, column=1, value="Account")
    ws_is.cell(row=r, column=2, value="12/31/2025")
    ws_is.cell(row=r, column=3, value=end_label)
    ws_is.cell(row=r, column=4, value="Change")
    _apply_header_row(ws_is, r, 4)
    r += 1

    for acct in ["Rental Income", "Interest Expense", "Accounting & Tax Fees",
                  "Bank Fees", "Depreciation Expense"]:
        beg = baseline_is.get(acct, 0)
        end = is_accounts.get(acct, 0)
        ws_is.cell(row=r, column=1, value=acct)
        _set_num_format(ws_is, r, 2, beg)
        _set_num_format(ws_is, r, 3, end)
        _set_num_format(ws_is, r, 4, end - beg)
        _apply_data_row(ws_is, r, 4, alt=(r % 2 == 0))
        r += 1

    ws_is.column_dimensions["A"].width = 30
    for c in ["B", "C", "D"]:
        ws_is.column_dimensions[c].width = 18

    # --- AJEs ---
    ws_aje = wb.create_sheet("Journal Entries")
    r = _write_title(ws_aje, 1, "{} | Journal Entries".format(fund_name), end_label, 4)
    r += 1
    ws_aje.cell(row=r, column=1, value="Date")
    ws_aje.cell(row=r, column=2, value="GL Account")
    ws_aje.cell(row=r, column=3, value="Debit")
    ws_aje.cell(row=r, column=4, value="Credit")
    _apply_header_row(ws_aje, r, 4)
    r += 1

    for entry in ajes:
        ws_aje.cell(row=r, column=1, value=entry["date"].strftime("%m/%d/%Y")).font = BOLD_FONT
        ws_aje.cell(row=r, column=2, value=entry["description"]).font = BOLD_FONT
        _apply_data_row(ws_aje, r, 4, bold=True)
        r += 1
        for acct, amt in entry["debits"].items():
            ws_aje.cell(row=r, column=2, value="  {}".format(acct)).font = LABEL_FONT
            _set_num_format(ws_aje, r, 3, amt)
            _apply_data_row(ws_aje, r, 4, alt=(r % 2 == 0))
            r += 1
        for acct, amt in entry["credits"].items():
            ws_aje.cell(row=r, column=2, value="      {}".format(acct)).font = LABEL_FONT
            _set_num_format(ws_aje, r, 4, amt)
            _apply_data_row(ws_aje, r, 4, alt=(r % 2 == 0))
            r += 1
        r += 1

    ws_aje.column_dimensions["A"].width = 14
    ws_aje.column_dimensions["B"].width = 38
    ws_aje.column_dimensions["C"].width = 18
    ws_aje.column_dimensions["D"].width = 18

    # --- Bank Activity ---
    ws_bank = wb.create_sheet("Bank Activity")
    r = _write_title(ws_bank, 1, "{} | Bank Activity".format(fund_name), end_label, 5)
    r += 1
    for i, h in enumerate(["Date", "Description", "Debit", "Credit", "Category"], 1):
        ws_bank.cell(row=r, column=i, value=h)
    _apply_header_row(ws_bank, r, 5)
    r += 1

    for t in txns:
        ws_bank.cell(row=r, column=1, value=t["post_date"][:10] if t["post_date"] else "")
        ws_bank.cell(row=r, column=2, value=t["description"])
        if t["debit"]:
            _set_num_format(ws_bank, r, 3, t["debit"])
        if t["credit"]:
            _set_num_format(ws_bank, r, 4, t["credit"])
        ws_bank.cell(row=r, column=5, value=t.get("details") or t.get("category") or "")
        _apply_data_row(ws_bank, r, 5, alt=(r % 2 == 0))
        r += 1

    ws_bank.column_dimensions["A"].width = 14
    ws_bank.column_dimensions["B"].width = 50
    ws_bank.column_dimensions["C"].width = 16
    ws_bank.column_dimensions["D"].width = 16
    ws_bank.column_dimensions["E"].width = 30

    wb.save(buffer)
    buffer.seek(0)
    return buffer
